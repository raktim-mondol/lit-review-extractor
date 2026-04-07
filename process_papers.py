"""
Literature Review Processor
Scans all .md files from the paper_in_markdown/ folder, processes each via
the Dashscope API, and writes structured results to an Excel file in result/.

Usage:
    1. Fill in DASHSCOPE_API_KEY in the .env file next to this script.
    2. Edit columns_config.json to define the columns you want extracted.
    3. python process_papers.py

Column configuration (columns_config.json):
    Each entry is a JSON object with:
        "group"        – (optional) Excel tier-1 header; consecutive columns with the
                         same group are merged on row 1. Defaults to "General".
        "column_name"  – tier-2 header (row 2) / Excel column label
        "field_key"    – the JSON key the model must return (no spaces, snake_case)
        "description"  – the instruction sent to the model for that field
        "width"        – (optional) Excel column width, defaults to 25

    The fixed columns "Serial No.", "File Name", "Status", and "Error" are
    always present and cannot be removed via the config. They use tier-1 groups
    "Identification" (prefix) and "Processing" (suffix).

    Excel layout: row 1 = merged group labels; row 2 = column names; data from row 3.
    If you have an old workbook with a single header row, delete it and re-run so
    headers are recreated.

Notes:
    - Creates the result/ folder automatically if it does not exist.
    - Creates a blank Excel file with headers on first run.
    - Appends one row per paper after each API call.
    - Checkpoint file tracks progress so you can safely stop and resume.
    - Use --force <serial> ... to reprocess specific papers.
    - Use --recreate-excel to replace the workbook with a fresh grouped-header
      template (previous file is moved to result/literature_review_backup_<timestamp>.xlsx).
      Does not call the API; run again without the flag to process papers.
"""

import argparse
import json
import os
import shutil
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load .env from the same directory as this script
load_dotenv(Path(__file__).resolve().parent / ".env")

# ── Configuration ──────────────────────────────────────────────────────────────
DASHSCOPE_API_KEY = os.environ.get("DASHSCOPE_API_KEY", "")
DASHSCOPE_BASE_URL = os.environ.get(
    "DASHSCOPE_BASE_URL", "https://coding-intl.dashscope.aliyuncs.com/v1"
)
MODEL = os.environ.get("DASHSCOPE_MODEL", "qwen3-max-2026-01-23")
ALT_DASHSCOPE_API_KEY = os.environ.get("ALT_DASHSCOPE_API_KEY", "")
ALT_DASHSCOPE_BASE_URL = os.environ.get("ALT_DASHSCOPE_BASE_URL", "")
ALT_DASHSCOPE_MODEL = os.environ.get("ALT_DASHSCOPE_MODEL", "")
ALT_DASHSCOPE_MODELS = os.environ.get("ALT_DASHSCOPE_MODELS", "")
ALT_EACH_MODEL_BATCH = int(
    os.environ.get("ALT_EACH_MODEL_BATCH", os.environ.get("ALT_FIRST_BATCH", "3"))
)
PRIMARY_NEXT_BATCH = int(os.environ.get("PRIMARY_NEXT_BATCH", "5"))
RETRY_ATTEMPTS = 3
RETRY_DELAY = 10  # seconds between retries
SLEEP_BETWEEN_CALLS = 2  # seconds between successful API calls
MAX_PAPER_CHARS = int(os.environ.get("MAX_PAPER_CHARS", "180000"))

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
MARKDOWN_DIR = SCRIPT_DIR / "paper_in_markdown"
RESULT_DIR = SCRIPT_DIR / "result"
EXCEL_OUTPUT = RESULT_DIR / "literature_review.xlsx"
JSON_OUTPUT_DIR = RESULT_DIR / "json_outputs"
CHECKPOINT_FILE = RESULT_DIR / "progress_checkpoint.json"
RUNTIME_STATUS_FILE = RESULT_DIR / "runtime_status.json"
GUIDELINE_PATH = SCRIPT_DIR / "guideline.md"
COLUMNS_CONFIG_PATH = SCRIPT_DIR / "columns_config.json"

# ── Default system prompt (used when guideline.md is absent) ──────────────────
DEFAULT_SYSTEM_PROMPT = (
    "You are an expert research assistant specialising in AI fairness and bias "
    "mitigation in computational/digital pathology. Your task is to extract "
    "structured information from academic papers and return it as a valid JSON "
    "object. Be precise, concise, and use only information from the provided text. "
    "If a field is not present in the paper, use exactly: Not Reported (NR)."
)

# ── Fixed columns (always present, not configurable) ──────────────────────────
# Prefix columns come before the user-defined columns.
# Suffix columns come after.
FIXED_PREFIX = [
    ("Serial No.", 10, "Identification"),
    ("File Name", 30, "Identification"),
]
FIXED_SUFFIX = [("Status", 12, "Processing"), ("Error", 30, "Processing")]


# ── Helpers ────────────────────────────────────────────────────────────────────


def load_columns_config() -> list[dict]:
    """
    Load and validate columns_config.json.

    Each entry must have:
        column_name  (str) – Excel header label
        field_key    (str) – JSON key the model returns; must be unique, snake_case
        description  (str) – instruction sent to the model for this field
        width        (int, optional) – Excel column width, default 25
        group        (str, optional) – tier-1 Excel group header, default "General"
    """
    if not COLUMNS_CONFIG_PATH.exists():
        print(f"[ERROR] columns_config.json not found: {COLUMNS_CONFIG_PATH}")
        print("        Create the file or copy the default from the repository.")
        sys.exit(1)

    try:
        config = json.loads(COLUMNS_CONFIG_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"[ERROR] columns_config.json is not valid JSON: {e}")
        sys.exit(1)

    if not isinstance(config, list) or len(config) == 0:
        print("[ERROR] columns_config.json must be a non-empty JSON array.")
        sys.exit(1)

    required_keys = {"column_name", "field_key", "description"}
    seen_keys: set[str] = set()
    for i, entry in enumerate(config):
        missing = required_keys - entry.keys()
        if missing:
            print(f"[ERROR] columns_config.json entry #{i + 1} is missing: {missing}")
            sys.exit(1)
        fk = entry["field_key"]
        if fk in seen_keys:
            print(f"[ERROR] columns_config.json has duplicate field_key: '{fk}'")
            sys.exit(1)
        seen_keys.add(fk)
        # Apply default width and Excel group
        entry.setdefault("width", 25)
        entry.setdefault("group", "General")

    return config


def build_excel_layout(columns_config: list[dict]) -> list[tuple[str, int, str]]:
    """(column_name, width, group) for every Excel column, fixed + user + fixed."""
    layout: list[tuple[str, int, str]] = []
    layout.extend(FIXED_PREFIX)
    layout.extend((c["column_name"], c["width"], c["group"]) for c in columns_config)
    layout.extend(FIXED_SUFFIX)
    return layout


def get_first_data_row(ws) -> int:
    """Row index where data starts (3 if tier-1/tier-2 headers exist, else 2)."""
    v = ws.cell(row=2, column=1).value
    if v is not None and str(v).strip() == "Serial No.":
        return 3
    return 2


def load_system_prompt() -> str:
    """Return guideline.md content when available, otherwise the built-in default."""
    if GUIDELINE_PATH.exists():
        return GUIDELINE_PATH.read_text(encoding="utf-8")
    return DEFAULT_SYSTEM_PROMPT


def load_checkpoint() -> dict:
    if CHECKPOINT_FILE.exists():
        checkpoint = json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
        checkpoint.setdefault("completed", [])
        checkpoint.setdefault("failed", [])
        checkpoint.setdefault("completed_files", [])
        checkpoint.setdefault("failed_files", [])
        return checkpoint
    return {"completed": [], "failed": [], "completed_files": [], "failed_files": []}


def save_checkpoint(checkpoint: dict) -> None:
    # Keep checkpoint lists deterministic and duplicate-free for stable resumes.
    checkpoint["completed"] = sorted({str(x) for x in checkpoint.get("completed", [])})
    checkpoint["failed"] = sorted({str(x) for x in checkpoint.get("failed", [])})
    checkpoint["completed_files"] = sorted(
        {str(x) for x in checkpoint.get("completed_files", [])}
    )
    checkpoint["failed_files"] = sorted(
        {str(x) for x in checkpoint.get("failed_files", [])}
    )
    CHECKPOINT_FILE.write_text(json.dumps(checkpoint, indent=2), encoding="utf-8")


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def update_runtime_status(**fields) -> None:
    """Persist a small live status payload for dashboard polling."""
    payload = {}
    if RUNTIME_STATUS_FILE.exists():
        try:
            payload = json.loads(RUNTIME_STATUS_FILE.read_text(encoding="utf-8"))
        except Exception:
            payload = {}
    payload.update(fields)
    payload["updated_at"] = utc_now_iso()
    RUNTIME_STATUS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def update_timing_stats(file_duration_sec: float) -> None:
    """
    Keep lightweight rolling timing stats for ETA estimation.
    Uses only current run data and ignores skipped files.
    """
    payload = {}
    if RUNTIME_STATUS_FILE.exists():
        try:
            payload = json.loads(RUNTIME_STATUS_FILE.read_text(encoding="utf-8"))
        except Exception:
            payload = {}

    processed = int(payload.get("processed_in_run", 0))
    avg_sec = float(payload.get("avg_file_seconds", 0.0))
    new_processed = processed + 1
    new_avg = (
        ((avg_sec * processed) + file_duration_sec) / new_processed
        if new_processed > 0
        else file_duration_sec
    )
    payload["processed_in_run"] = new_processed
    payload["avg_file_seconds"] = round(new_avg, 2)
    payload["last_file_seconds"] = round(file_duration_sec, 2)
    payload["updated_at"] = utc_now_iso()
    RUNTIME_STATUS_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def collect_markdown_files() -> list[Path]:
    """Return all .md files in MARKDOWN_DIR sorted by name."""
    if not MARKDOWN_DIR.exists():
        print(f"[ERROR] Markdown folder not found: {MARKDOWN_DIR}")
        sys.exit(1)
    files = sorted(MARKDOWN_DIR.glob("*.md"))
    if not files:
        print(f"[ERROR] No .md files found in {MARKDOWN_DIR}")
        sys.exit(1)
    return files


def write_excel_workbook_shell(columns_config: list[dict]) -> None:
    """Write EXCEL_OUTPUT with grouped tier-1 + tier-2 headers only (overwrites path)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature Review"

    group_fill = PatternFill(
        start_color="2E5090", end_color="2E5090", fill_type="solid"
    )
    sub_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    group_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    layout = build_excel_layout(columns_config)
    ncols = len(layout)

    # Row 1: merged group labels
    i = 0
    while i < ncols:
        j = i
        group_name = layout[i][2]
        while j < ncols and layout[j][2] == group_name:
            j += 1
        start_c = i + 1
        end_c = j
        if end_c > start_c:
            ws.merge_cells(
                start_row=1, start_column=start_c, end_row=1, end_column=end_c
            )
        c1 = ws.cell(row=1, column=start_c, value=group_name)
        c1.fill = group_fill
        c1.font = header_font
        c1.alignment = group_align
        c1.border = border
        for cc in range(start_c + 1, end_c + 1):
            cn = ws.cell(row=1, column=cc)
            cn.fill = group_fill
            cn.border = border
        i = j

    ws.row_dimensions[1].height = 28

    # Row 2: column titles + widths
    ws.row_dimensions[2].height = 36
    for col_idx, (col_name, col_width, _group) in enumerate(layout, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = sub_fill
        cell.font = header_font
        cell.alignment = sub_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A3"
    wb.save(EXCEL_OUTPUT)
    print(f"[INIT] Created blank Excel: {EXCEL_OUTPUT}")
    print(f"[INIT] Groups: {list(dict.fromkeys(x[2] for x in layout))}")
    print(f"[INIT] Columns: {[x[0] for x in layout]}")


def create_excel_if_missing(columns_config: list[dict]) -> None:
    """Create a blank Excel file if it does not exist yet."""
    if EXCEL_OUTPUT.exists():
        return
    write_excel_workbook_shell(columns_config)


def append_row_to_excel(row_data: list) -> None:
    """Append a single data row to the Excel file (after tier-1/tier-2 headers)."""
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    ws = wb.active

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(vertical="top", wrap_text=True)

    next_row = ws.max_row + 1
    fill_color = "F2F7FF" if next_row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(
        start_color=fill_color, end_color=fill_color, fill_type="solid"
    )

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(
            row=next_row, column=col_idx, value=str(value) if value is not None else ""
        )
        cell.alignment = data_align
        cell.border = border
        cell.fill = row_fill

    ws.row_dimensions[next_row].height = 60
    wb.save(EXCEL_OUTPUT)


def build_user_prompt(paper_content: str, columns_config: list[dict]) -> str:
    """Build the extraction prompt from the user-defined column config."""
    schema_items = [
        f'  "{col["field_key"]}": "{col["description"]}"' for col in columns_config
    ]
    schema = "{\n" + ",\n".join(schema_items) + "\n}"

    field_list = ", ".join(f'"{col["field_key"]}"' for col in columns_config)

    trimmed_content = paper_content
    trim_note = ""
    if len(paper_content) > MAX_PAPER_CHARS:
        trimmed_content = paper_content[:MAX_PAPER_CHARS]
        trim_note = (
            f"\n- Input text was truncated to first {MAX_PAPER_CHARS} characters "
            "to stay within model context limits."
        )

    return f"""Below is the full text of a research paper. Extract the required information and return ONLY a valid JSON object (no markdown, no explanation, no extra text) with exactly these keys:

{schema}

Rules:
- Return ONLY the JSON object — no code fences, no extra text before or after.
- Use ONLY information explicitly stated in the paper.
- If a field is not mentioned in the paper, use exactly: Not Reported (NR)
- All keys must be present: {field_list}
- Keep values concise and factual.{trim_note}

--- PAPER CONTENT START ---
{trimmed_content}
--- PAPER CONTENT END ---
"""


def call_api(client: OpenAI, model: str, system_prompt: str, user_prompt: str) -> str:
    """Call the API and return the assistant's text content."""
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        extra_body={"reasoning": {"enabled": True}},
    )
    return response.choices[0].message.content or ""


def parse_json_response(raw: str) -> dict:
    """Parse JSON from the model response, tolerating minor formatting issues."""
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start != -1 and end > start:
        raw = raw[start:end]
    return json.loads(raw)


def remove_excel_rows(serials: set) -> None:
    """Delete rows from the Excel file whose Serial No. matches any of the given serials."""
    if not EXCEL_OUTPUT.exists():
        return
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    ws = wb.active
    data_start = get_first_data_row(ws)
    rows_to_delete = [
        row_idx
        for row_idx in range(data_start, ws.max_row + 1)
        if str(ws.cell(row=row_idx, column=1).value) in serials
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)
    wb.save(EXCEL_OUTPUT)
    if rows_to_delete:
        print(
            f"[FORCE] Removed {len(rows_to_delete)} existing Excel row(s) for serials: {serials}"
        )


def parse_args():
    parser = argparse.ArgumentParser(description="Literature Review Processor")
    parser.add_argument(
        "--force",
        nargs="+",
        metavar="SERIAL",
        help="Force re-process specific serial numbers even if already completed or failed. "
        "Example: --force 1 3 5",
    )
    parser.add_argument(
        "--recreate-excel",
        action="store_true",
        help="Replace literature_review.xlsx with a fresh header-only workbook from columns_config.json. "
        "Renames any existing file to literature_review_backup_<timestamp>.xlsx. Does not call the API.",
    )
    return parser.parse_args()


def select_route(
    file_counter: int,
    alt_enabled: bool,
    alt_models: list[str],
    alt_batch_per_model: int,
    primary_batch: int,
) -> tuple[str, str | None, str]:
    """
    Route current non-skipped file index to:
    - One ALT model (each repeated N times), then
    - PRIMARY model for configured batch, then repeat.
    Returns: (route_kind, selected_alt_model_or_none, route_label)
    """
    if not alt_enabled or not alt_models or alt_batch_per_model <= 0:
        return ("primary", None, "PRIMARY")

    alt_total = len(alt_models) * alt_batch_per_model
    cycle = alt_total + max(primary_batch, 1)
    position = file_counter % cycle
    if position < alt_total:
        alt_idx = position // alt_batch_per_model
        model = alt_models[alt_idx]
        return ("alt", model, f"ALT-{alt_idx + 1}/{len(alt_models)}")
    return ("primary", None, "PRIMARY")


def detect_provider(base_url: str) -> str:
    url = (base_url or "").lower()
    if "openrouter.ai" in url:
        return "openrouter"
    if "dashscope" in url or "aliyuncs.com" in url:
        return "dashscope"
    return "custom"


# ── Main ───────────────────────────────────────────────────────────────────────


def main():
    args = parse_args()
    force_serials: set = {str(s) for s in args.force} if args.force else set()

    # Ensure output directories exist
    RESULT_DIR.mkdir(exist_ok=True)
    JSON_OUTPUT_DIR.mkdir(exist_ok=True)

    # ── Load column config ────────────────────────────────────────────────────
    columns_config = load_columns_config()

    if args.recreate_excel:
        if EXCEL_OUTPUT.exists():
            ts = time.strftime("%Y%m%d_%H%M%S")
            backup = RESULT_DIR / f"literature_review_backup_{ts}.xlsx"
            shutil.move(EXCEL_OUTPUT, backup)
            print(f"[RECREATE] Existing workbook moved to {backup}")
        write_excel_workbook_shell(columns_config)
        print(
            "[RECREATE] Header-only workbook is ready. Run without --recreate-excel to process papers."
        )
        print(
            "[RECREATE] progress_checkpoint.json was not modified; reset it or use --force if needed."
        )
        return

    # ── Startup checks ────────────────────────────────────────────────────────
    if not DASHSCOPE_API_KEY:
        print("[ERROR] DASHSCOPE_API_KEY is not set.")
        print("        Add it to the .env file next to this script:")
        print("        DASHSCOPE_API_KEY=your_api_key_here")
        sys.exit(1)

    # Collect markdown files
    md_files = collect_markdown_files()
    total = len(md_files)

    print(f"[INFO] Script dir    : {SCRIPT_DIR}")
    print(f"[INFO] Markdown dir  : {MARKDOWN_DIR}")
    print(f"[INFO] Columns config: {COLUMNS_CONFIG_PATH}")
    print(f"[INFO] Excel output  : {EXCEL_OUTPUT}")
    print(f"[INFO] JSON outputs  : {JSON_OUTPUT_DIR}")
    print(f"[INFO] Papers found  : {total}")
    print(f"[INFO] User columns  : {[c['column_name'] for c in columns_config]}")
    update_runtime_status(
        state="running",
        total_papers=total,
        message="Run initialized.",
        started_at=utc_now_iso(),
        processed_in_run=0,
        avg_file_seconds=0.0,
        last_file_seconds=0.0,
    )

    primary_client = OpenAI(base_url=DASHSCOPE_BASE_URL, api_key=DASHSCOPE_API_KEY)
    raw_alt_models = [
        m.strip() for m in ALT_DASHSCOPE_MODELS.split(",") if m.strip()
    ]
    alt_models = raw_alt_models if raw_alt_models else ([ALT_DASHSCOPE_MODEL] if ALT_DASHSCOPE_MODEL else [])

    alt_enabled = bool(
        ALT_DASHSCOPE_API_KEY and ALT_DASHSCOPE_BASE_URL and len(alt_models) > 0
    )
    alt_client = (
        OpenAI(base_url=ALT_DASHSCOPE_BASE_URL, api_key=ALT_DASHSCOPE_API_KEY)
        if alt_enabled
        else None
    )
    if alt_enabled:
        print(
            f"[INFO] Model routing : ALT models {alt_models} each x{ALT_EACH_MODEL_BATCH} "
            f"-> PRIMARY({MODEL}) x{PRIMARY_NEXT_BATCH} (repeating)"
        )
    else:
        print(f"[INFO] Model routing : PRIMARY only ({MODEL})")
    system_prompt = load_system_prompt()
    checkpoint = load_checkpoint()
    create_excel_if_missing(columns_config)

    # Backward compatibility: map legacy serial-based progress to filenames.
    serial_to_filename = {str(i): p.name for i, p in enumerate(md_files, start=1)}
    completed_files = set(checkpoint.get("completed_files", []))
    failed_files = set(checkpoint.get("failed_files", []))
    completed_files.update(
        serial_to_filename[s]
        for s in checkpoint.get("completed", [])
        if s in serial_to_filename
    )
    failed_files.update(
        serial_to_filename[s]
        for s in checkpoint.get("failed", [])
        if s in serial_to_filename
    )
    checkpoint["completed_files"] = sorted(completed_files)
    checkpoint["failed_files"] = sorted(failed_files)

    # Apply --force: remove forced serials from checkpoint and Excel
    if force_serials:
        print(f"[FORCE] Forcing re-process of serials: {sorted(force_serials)}")
        checkpoint["completed"] = [
            s for s in checkpoint["completed"] if s not in force_serials
        ]
        checkpoint["failed"] = [
            s for s in checkpoint["failed"] if s not in force_serials
        ]
        forced_filenames = {
            serial_to_filename[s] for s in force_serials if s in serial_to_filename
        }
        checkpoint["completed_files"] = [
            f for f in checkpoint["completed_files"] if f not in forced_filenames
        ]
        checkpoint["failed_files"] = [
            f for f in checkpoint["failed_files"] if f not in forced_filenames
        ]
        save_checkpoint(checkpoint)
        remove_excel_rows(force_serials)

    print(
        f"[INFO] Already processed: {len(checkpoint['completed'])} | Failed: {len(checkpoint['failed'])}"
    )

    n_user_cols = len(columns_config)
    route_counter = 0

    for serial_int, md_path in enumerate(md_files, start=1):
        serial = str(serial_int)
        filename = md_path.name

        # When --force is given, only process the specified serials
        if force_serials and serial not in force_serials:
            continue

        # Skip already completed (unless forced). Prefer filename-based resume.
        if filename in checkpoint["completed_files"] or serial in checkpoint["completed"]:
            print(f"[SKIP] #{serial}/{total} — {filename} (already processed)")
            continue

        print(f"\n[PROCESSING] #{serial}/{total} — {filename}")
        file_start = time.time()
        route, alt_model, route_tag = select_route(
            file_counter=route_counter,
            alt_enabled=alt_enabled,
            alt_models=alt_models,
            alt_batch_per_model=ALT_EACH_MODEL_BATCH,
            primary_batch=PRIMARY_NEXT_BATCH,
        )
        route_counter += 1
        selected_model = alt_model if route == "alt" and alt_model else MODEL
        selected_client = alt_client if route == "alt" else primary_client
        selected_base_url = ALT_DASHSCOPE_BASE_URL if route == "alt" else DASHSCOPE_BASE_URL
        selected_provider = detect_provider(selected_base_url)
        update_runtime_status(
            state="running",
            current_serial=serial,
            current_file=filename,
            model=selected_model,
            provider=selected_provider,
            route=route_tag,
            message=f"Processing {filename} [{selected_provider}/{selected_model}]",
        )

        try:
            paper_content = md_path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            err_msg = f"[ERROR reading file: {e}]"
            print(f"  [WARN] {err_msg}")
            row = (
                [serial, filename]
                + ["Not Reported (NR)"] * n_user_cols
                + ["FAILED", err_msg]
            )
            append_row_to_excel(row)
            checkpoint["failed"].append(serial)
            checkpoint["failed_files"].append(filename)
            save_checkpoint(checkpoint)
            update_timing_stats(time.time() - file_start)
            update_runtime_status(
                state="running",
                current_serial=serial,
                current_file=filename,
                message=f"Failed to read file: {filename}",
                last_result="failed",
                last_error=err_msg,
            )
            continue

        user_prompt = build_user_prompt(paper_content, columns_config)

        # Retry loop
        success = False
        last_error = ""
        extracted = {}

        for attempt in range(1, RETRY_ATTEMPTS + 1):
            try:
                if attempt > 1:
                    print(f"  [API ] Retry attempt {attempt}/{RETRY_ATTEMPTS}...")
                update_runtime_status(
                    state="running",
                    current_serial=serial,
                    current_file=filename,
                    model=selected_model,
                    provider=selected_provider,
                    route=route_tag,
                    attempt=attempt,
                    message=(
                        f"Calling {selected_provider}/{selected_model} for {filename} "
                        f"(attempt {attempt}/{RETRY_ATTEMPTS})"
                    ),
                )
                raw_response = call_api(
                    selected_client, selected_model, system_prompt, user_prompt
                )
                extracted = parse_json_response(raw_response)
                success = True
                break
            except json.JSONDecodeError as e:
                last_error = f"JSON parse error: {e} | Raw: {raw_response[:200]}"
                print(f"  [WARN] {last_error}")
            except Exception as e:
                last_error = str(e)
                print(f"  [WARN] API error: {last_error}")

            if attempt < RETRY_ATTEMPTS:
                print(f"  [WAIT] Retrying in {RETRY_DELAY}s...")
                time.sleep(RETRY_DELAY)

        if success:
            # Save individual JSON output to result/json_outputs/
            json_path = JSON_OUTPUT_DIR / f"{serial}_{md_path.stem}.json"
            json_path.write_text(
                json.dumps(
                    {"serial": serial, "filename": filename, **extracted},
                    indent=2,
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            # Build row: prefix + one value per user column + suffix
            user_values = [
                extracted.get(col["field_key"], "Not Reported (NR)")
                for col in columns_config
            ]
            row = [serial, filename] + user_values + ["SUCCESS", ""]
            append_row_to_excel(row)
            checkpoint["completed"].append(serial)
            checkpoint["completed_files"].append(filename)
            checkpoint["failed"] = [s for s in checkpoint["failed"] if s != serial]
            checkpoint["failed_files"] = [
                f for f in checkpoint["failed_files"] if f != filename
            ]
            save_checkpoint(checkpoint)
            update_timing_stats(time.time() - file_start)
            update_runtime_status(
                state="running",
                current_serial=serial,
                current_file=filename,
                message=f"Completed {filename}",
                last_result="success",
                last_error="",
            )
            print(f"  [OK  ] Saved to Excel and JSON.")
        else:
            row = (
                [serial, filename]
                + ["Not Reported (NR)"] * n_user_cols
                + ["FAILED", last_error[:500]]
            )
            append_row_to_excel(row)
            checkpoint["failed"].append(serial)
            checkpoint["failed_files"].append(filename)
            save_checkpoint(checkpoint)
            update_timing_stats(time.time() - file_start)
            update_runtime_status(
                state="running",
                current_serial=serial,
                current_file=filename,
                message=f"Failed {filename}",
                last_result="failed",
                last_error=last_error[:500],
            )
            print(f"  [FAIL] Logged failure.")

        time.sleep(SLEEP_BETWEEN_CALLS)

    completed = len(checkpoint["completed"])
    failed = len(checkpoint["failed"])
    print(f"\n[DONE] Processed {completed} papers successfully, {failed} failed.")
    print(f"       Excel : {EXCEL_OUTPUT}")
    print(f"       JSONs : {JSON_OUTPUT_DIR}")
    if checkpoint["failed"]:
        print(f"       Failed serial numbers: {checkpoint['failed']}")
    update_runtime_status(
        state="idle",
        current_serial=None,
        current_file=None,
        message="Run finished.",
        finished_at=utc_now_iso(),
        last_result="done",
    )


if __name__ == "__main__":
    main()
